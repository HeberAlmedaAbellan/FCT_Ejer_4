import os
from datetime import datetime
import openpyxl
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait


class Licitaciones:

    def __init__(self):
        self.PATH = "C:\\Program Files (x86)\\chromedriver.exe"
        self.driver = webdriver.Chrome(self.PATH)
        self.num_cols = ""
        self.num_rows = ""
        self.titles = []
        self.expediente = []
        self.tipo_contrato = []
        self.estado = []
        self.importe = []
        self.fechas = []
        self.organo_contratacion = []
        self.ruta_expedientes="./Expedientes"

    def nav(self):
        self.driver.maximize_window()
        self.driver.get("https://contrataciondelestado.es/")
        VIEWNS = "viewns_Z7_AVEQAI930OBRD02JPMTPG21004_"

        search = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located(
                (By.LINK_TEXT, "Licitaciones")
            )
        )
        search.click()
        search = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located(
                (By.LINK_TEXT, "Búsqueda")
            )
        )
        search.click()

        search = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located(
                (By.ID, f"{VIEWNS}:form1:logoFormularioBusqueda")
            )
        )
        search.click()

        search = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located(
                (By.ID, f"{VIEWNS}:form1:linkBusquedaAvanzada")
            )
        )
        search.click()

        search = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located(
                (By.ID, f"{VIEWNS}:form1:menu111MAQ")
            )
        )
        search.click()

        search = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located(
                (By.XPATH, f"//select[@id='{VIEWNS}:form1:menu111MAQ']/option[@value='1']")
            )
        )
        search.click()

        search = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located(
                (By.ID, f"{VIEWNS}:form1:tipoSistemaContratacion")
            )
        )
        search.click()

        search = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located(
                (By.XPATH, f"//select[@id='{VIEWNS}:form1:tipoSistemaContratacion']"
                           "/option[@value='3']")
            )
        )
        search.click()

        search = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located(
                (By.ID, f"{VIEWNS}:form1:s_detLicitacionMAQ")
            )
        )
        search.click()

        search = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located(
                (By.XPATH, f"//select[@id='{VIEWNS}:form1:s_detLicitacionMAQ']"
                           "/option[@value='DOC_CN']")
            )
        )
        search.click()

        search = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located(
                (By.ID, f"{VIEWNS}:form1:menuTipoContMAQ")
            )
        )
        search.click()

        search = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located(
                (By.XPATH, f"//select[@id='{VIEWNS}:form1:menuTipoContMAQ']"
                           "/option[@value='2']")
            )
        )
        search.click()

        search = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located(
                (By.ID, f"{VIEWNS}:form1:menuSubtipoMAQ")
            )
        )
        search.click()

        search = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located(
                (By.XPATH, f"//select[@id='{VIEWNS}:form1:menuSubtipoMAQ']"
                           "/option[@value='7']")
            )
        )
        search.click()

        search = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located(
                (By.ID, f"{VIEWNS}:form1:button1")
            )
        )
        search.click()

    def get_num_rows(self):

        WebDriverWait(self.driver, 10).until(EC.presence_of_element_located(
            (By.XPATH, "//*[@id='myTablaBusquedaCustom']/thead/tr/th[1]/div/span")))

        self.num_rows = len(self.driver.find_elements_by_xpath("//*[@id='myTablaBusquedaCustom']/tbody/tr"))

    def get_num_cols(self):

        WebDriverWait(self.driver, 10).until(EC.presence_of_element_located(
            (By.XPATH, "//*[@id='myTablaBusquedaCustom']/thead/tr/th[1]/div/span")))

        self.num_cols = len(self.driver.find_elements_by_xpath("//*[@id='myTablaBusquedaCustom']/tbody/tr[2]/td"))

    def get_data(self):

        titles =[]
        TABLE_ID = "myTablaBusquedaCustom"

        for i in range(self.num_cols):
            t = \
                self.driver.find_element(By.XPATH, f"//*[@id='{TABLE_ID}']/thead/tr/th[{i+1}]/div/span").text
            titles.append(t)

        self.titles = titles

        exp = []

        for i in range(self.num_rows):
            d = \
                self.driver.find_element(
                    By.XPATH, f"// *[ @ id = '{TABLE_ID}'] / tbody / tr[{i+1}] / td[1] / div[2]").text
            exp.append(d)

        self.expediente = exp

        tipc = []

        for i in range(self.num_rows):
            d = \
                self.driver.find_element(
                    By.XPATH, f"// *[ @ id = '{TABLE_ID}'] / tbody / tr[{i + 1}] / td[2] / div[2]").text
            tipc.append(d)

        self.tipo_contrato = tipc

        estado = []

        for i in range(self.num_rows):
            d = \
                self.driver.find_element(
                    By.XPATH, f"// *[ @ id = '{TABLE_ID}'] / tbody / tr[{i + 1}] / td[3]").text
            estado.append(d)

        self.estado = estado

        importe = []

        for i in range(self.num_rows):
            d = \
                self.driver.find_element(
                    By.XPATH, f"// *[ @ id = '{TABLE_ID}'] / tbody / tr[{i + 1}] / td[4]").text
            importe.append(d)

        self.importe = importe

        fechas = []

        for i in range(self.num_rows):
            if i != 2:
                d = \
                    self.driver.find_element(
                        By.XPATH,
                        f"// *[ @ id = '{TABLE_ID}'] / tbody / tr[{i + 1}] / td[5] / div[1] /span").text
                d += " " + \
                    self.driver.find_element(
                        By.XPATH,
                        f"// *[ @ id = '{TABLE_ID}'] / tbody / tr[{i + 1}] / td[5] / div[2] / span[1]").text
                d += " " + \
                    self.driver.find_element(
                        By.XPATH,
                        f"// *[ @ id = '{TABLE_ID}'] / tbody / tr[{i + 1}] / td[5] / div[2] / span[2]").text
                d += " " + \
                    self.driver.find_element(
                        By.XPATH,
                        f"// *[ @ id = '{TABLE_ID}'] / tbody / tr[{i + 1}] / td[5] / div[3] / span[1]").text
                d += " " + \
                    self.driver.find_element(
                        By.XPATH,
                        f"// *[ @ id = '{TABLE_ID}'] / tbody / tr[{i + 1}] / td[5] / div[3] / span[2]").text
                fechas.append(d)
            else:
                d = \
                    self.driver.find_element(
                        By.XPATH,
                        f"//*[@id='{TABLE_ID}']/tbody/tr[3]/td[5]/div/span[1]").text
                d += " " + \
                     self.driver.find_element(
                         By.XPATH,
                         f"//*[@id='{TABLE_ID}']/tbody/tr[3]/td[5]/div/span[2]").text
                fechas.append(d)

        self.fechas = fechas

        org_contratacion = []

        for i in range(self.num_rows):
            d = \
                self.driver.find_element(
                    By.XPATH, f"// *[ @ id = '{TABLE_ID}'] / tbody / tr[{i + 1}] / td[6] / a").text
            org_contratacion.append(d)

        self.organo_contratacion = org_contratacion

    def data_to_excel(self):

        my_wb = openpyxl.Workbook()
        my_sheet = my_wb.active

        # cabeceras
        for i in range(self.num_cols):
            c1 = my_sheet.cell(row=1, column=i+1)
            c1.value = self.titles[i]

        # contenido primera columna
        for i in range(self.num_rows):
            c1 = my_sheet.cell(row=i+2, column=1)
            c1.value = self.expediente[i]

        # contenido segunda columna
        for i in range(self.num_rows):
            c1 = my_sheet.cell(row=i + 2, column=2)
            c1.value = self.tipo_contrato[i]

        # contenido tercera columna
        for i in range(self.num_rows):
            c1 = my_sheet.cell(row=i + 2, column=3)
            c1.value = self.estado[i]

        # contenido cuarta columna
        for i in range(self.num_rows):
            c1 = my_sheet.cell(row=i + 2, column=4)
            c1.value = self.importe[i]

        # contenido quinta columna
        for i in range(self.num_rows):
            c1 = my_sheet.cell(row=i + 2, column=5)
            c1.value = self.fechas[i]

        # contenido sexta columna
        for i in range(self.num_rows):
            c1 = my_sheet.cell(row=i + 2, column=6)
            c1.value = self.organo_contratacion[i]

        my_wb.save("licitaciones.xlsx")

    def move_file(self):
        os.rename("./licitaciones.xlsx", self.ruta_expedientes+"/licitaciones-" +
                  str(datetime.today().strftime('%A-%B-%d-%Y-%H-%M-%S'))+".xlsx")



